def update_xlsx(path = r'data.xlsx', id = 0):
   from datetime import datetime
   import openpyxl

   data = openpyxl.load_workbook(path)

   data_frame = data.active 

   for row in range(2, data_frame.max_row+1):
      # print(data_frame.cell(row, 1).value)
      if (id == str(data_frame.cell(row, 1).value)):
         data_frame.cell(row, 3).value = 'X'
         data_frame.cell(row, 4).value = datetime.now()
   
   data.save(path)
#update_xlsx(path = r'data.xlsx', id = message)

